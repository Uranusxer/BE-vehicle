from django.shortcuts import render
from django.http import HttpRequest
from item.models import Item
from parameter.models import Vehicle,Project,Site,Goods
from utils.utils_request import BAD_METHOD, request_failed, request_success
from utils.utils_require import CheckRequire, require
from utils.utils_time import get_timestamp
from django.core.paginator import Paginator
from user.models import User,get_user_from_request
from finance.models import Advance,Payment
from utils.constants import START,END
import json
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from django.db.models import Sum, F, FloatField, Value
import re
import io
from openpyxl.styles import Alignment, Font, Border, Side
from django.http import HttpRequest, HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side
import tempfile
import subprocess
import os
import cn2an
@CheckRequire
def advance(req:HttpRequest):
    # failure_response, user = get_user_from_request(req,'POST')
    # if failure_response:
    #     return failure_response
    body = json.loads(req.body.decode("utf-8"))
    vehicle_id = require(body,"vehicle_id","int",err_msg="Missing or error type of [vehicle_id]")
    pay_id = require(body,"pay_id","int",err_msg="Missing or error type of [pay_id]")
    amount = require(body,"amount","int",err_msg="Missing or error type of [amount]")
    advance_time = require(body,"advance_time","string",err_msg="Missing or error type of [advance_time]")
    try:
        note = require(body,"note","string",err_msg="Missing or error type of [note]")
    except:
        note = "无"
    Newadvance = Advance.objects.create(pay_id=pay_id,vehicle_id=vehicle_id,amount=amount,advance_time=advance_time,created_time=get_timestamp(),note=note)
    return request_success()

@CheckRequire
def del_advance(req:HttpRequest,advance_id):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    advance = Advance.objects.filter(id=advance_id,if_delete=False).first()
    if not advance:
        return request_failed(code=1,info="Advance does not exist",status_code=404)
    advance.if_delete=True
    advance.save()
    return request_success()

@CheckRequire
def advance_list(req:HttpRequest,per_page,page):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    driver = req.GET.get('driver',None)

    advances = Advance.objects.filter(if_delete=False)
    if driver is not None:
        vehicle_ids = Vehicle.objects.filter(driver=driver).values_list('id', flat=True)
        advances = advances.filter(vehicle_id__in=vehicle_ids)
   
    paginator = Paginator(advances, per_page)
    current_page = paginator.get_page(page)
    total_pages = paginator.num_pages
    return_data = [advance.serialize() for advance in current_page]
    return request_success({"advances":return_data,"total_pages":total_pages})

@CheckRequire
def change_advance(req:HttpRequest):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    body = json.loads(req.body.decode("utf-8"))
    advance_id = require(body,"advance_id","int",err_msg="Missing or error type of [advance_id]")
    advance = Advance.objects.filter(id=advance_id).first()
    if not advance:
        return request_failed(code=1,info="Advance does not exist",status_code=404)
    try:
        vehicle_id = require(body, "vehicle_id", "int", err_msg="Missing or error type of [vehicle_id]")
    except:
        vehicle_id = None
    try:
        pay_id = require(body, "pay_id", "int", err_msg="Missing or error type of [pay_id]")
    except:
        pay_id = None
    try:
        amount = require(body, "amount", "int", err_msg="Missing or error type of [amount]")
    except:
        amount = None
    try:
        advance_time = require(body, "advance_time", "string", err_msg="Missing or error type of [advance_time]")
    except:
        advance_time = None
    try:
        note = require(body, "note", "string", err_msg="Missing or error type of [note]")
    except:
        note = None
    if vehicle_id:
        advance.vehicle_id = vehicle_id
    if amount:
        advance.amount = amount
    if advance_time:
        advance.advance_time = advance_time
    if note:
        advance.note = note
    if pay_id:
        advance.pay_id = pay_id
    advance.save()
    return request_success()

@CheckRequire
def search4advance(req:HttpRequest,per_page,page):
    start_date = req.GET.get('start_date',None)
    end_date = req.GET.get('end_date',None)
    vehicle_id = req.GET.get('vehicle_id',None)
    advances = Advance.objects.filter(if_delete=False).order_by("-created_time")
    if vehicle_id:
        advances = advances.filter(vehicle_id=vehicle_id)
    if start_date:
        advances = advances.filter(advance_time__gte=start_date)
    if end_date:
        advances = advances.filter(advance_time__lte=end_date)
        
    paginator = Paginator(advances, per_page)
    current_page = paginator.get_page(page)
    total_pages = paginator.num_pages
    return_data = [advance.serialize() for advance in current_page]

    return request_success({"advances":return_data,"total_pages":total_pages})


@CheckRequire
def total_amount(req:HttpRequest):
    project_owner = req.GET.get('ownerName', None)
    project_id = req.GET.get('project_id', None)
    goods_id = req.GET.get('goods_id', None)
    startsite_id = req.GET.get('startsite_id',None)
    endsite_id = req.GET.get('endsite_id',None)
    start_date = req.GET.get('start_date',None)
    end_date = req.GET.get('end_date',None)
    items = Item.objects.filter(if_delete=False)
    if project_owner is not None: 
        project_ids = Project.objects.filter(owner=project_owner).values_list('id', flat=True)
        items = items.filter(project_id__in=project_ids)
    if project_id is not None:
        items = items.filter(project_id=project_id)
    if goods_id is not None:
        items = items.filter(goods_id=goods_id)
    if startsite_id is not None:
        items = items.filter(startsite_id=startsite_id)
    if endsite_id is not None:
        items = items.filter(endsite_id=endsite_id)

    if start_date is not None:
        items = items.filter(date__gte=start_date)
    if end_date is not None:
        items = items.filter(date__lte=end_date)
    total_amount = 0.0
    for item in items:
        total_amount += (item.contractorPrice - item.endPayment - item.driverPrice) * item.quantity + item.startSubsidy + item.endSubsidy
    return request_success({"total_amount":total_amount})



@CheckRequire
def driver_excel(req:HttpRequest):
    body = json.loads(req.body.decode("utf-8"))
    start_date = require(body, "start_date", "string", err_msg="Missing or error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing or error type of [end_date]")
    vehicle_id = require(body, "vehicle_id", "int", err_msg="Missing or error type of [vehicle_id]")
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")

    vehicle = Vehicle.objects.filter(id=vehicle_id).first()
    # 获取所有items
    items = Item.objects.filter(id__in=item_ids, if_delete=False)
    
    start_date = start_date.split('T')[0]
    end_date = end_date.split('T')[0]

    # 创建Excel工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运司机对账单"
    default_font = Font(size=12)
    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置列宽
    column_widths = {
        'A': 7,    # 序号
        'B': 15,     # 日期
        'C': 20,    # 运输起点
        'D': 20,    # 运输终点
        'E': 10,    # 品类
        'F': 8,    # 数量
        'G': 8,    # 单位
        'H': 10.75,     # 装车方式
        'I': 20,     # 给司机单价
        'J': 8,    # 金额
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width
    
    # 固定的表头信息

    sheet.merge_cells('A1:J1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运司机对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=16, bold=True)

    sheet.merge_cells('A2:B2')
    sheet.merge_cells('E2:G2')
    sheet.merge_cells('H2:J2')

    sheet['A2'] = f"车牌号：{vehicle.license}"
    sheet['C2'] = f"驾驶员：{vehicle.driver}"
    sheet['D2'] = f"电话：{vehicle.phone}"
    sheet['E2'] = f"起始日期：{start_date}"
    sheet['H2'] = f"截止日期：{end_date}"
    # 将第一行的字体设置为加粗，居中
    for row in sheet['A2:J2']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
    
    

    headers = ["序号", "日期", "运输起点", "运输终点", "品类", "数量", "单位", "装车方式", "给司机单价", "金额"]
    sheet.append(headers)
    current_row = sheet.max_row
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    total_amount = 0.0
    for idx, item in enumerate(items, start=1):
        start_site = Site.objects.filter(id=item.startsite_id).first()
        start_site_name = start_site.name if start_site else "无"
        end_site = Site.objects.filter(id=item.endsite_id).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        total_price = item.quantity * item.driverPrice
        row = [
            idx,
            item.date.split('T')[0],
            start_site_name,
            end_site_name,
            goods_name,
            item.quantity,
            item.unit,
            item.get_load_display(),
            item.driverPrice,
            total_price
        ]
        sheet.append(row)
        current_row = sheet.max_row
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        total_amount += total_price
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append([f"总 计 金 额：{total_amount}", "", "","",f"总 计 大 写 (金 额)：{total_cn}"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:D{current_row}')
    sheet.merge_cells(f'E{current_row}:J{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    sheet.row_dimensions[1].height = 25
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 18.5
    local_file_path = "/root/cheliangyunshu/BE-vehicle/test/driver.xlsx"
    workbook.save(local_file_path)
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment;filename=driver_excel.xlsx'
    return response


@CheckRequire
def driver_excel_pdf(req: HttpRequest):
    body = json.loads(req.body.decode("utf-8"))
    start_date = require(body, "start_date", "string", err_msg="Missing or error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing or error type of [end_date]")
    vehicle_id = require(body, "vehicle_id", "int", err_msg="Missing or error type of [vehicle_id]")
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    
    vehicle = Vehicle.objects.filter(id=vehicle_id).first()
    start_date = start_date.split('T')[0]
    end_date = end_date.split('T')[0]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运司机对账单"

    default_font = Font(size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    column_widths = {
        'A': 7, 'B': 15, 'C': 20, 'D': 20, 'E': 10,
        'F': 8, 'G': 8, 'H': 10.75, 'I': 20, 'J': 8
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    sheet.merge_cells('A1:J1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运司机对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=16, bold=True)
    title_cell.border = thin_border
    sheet['J1'].border = thin_border
    sheet.merge_cells('A2:B2')
    sheet.merge_cells('E2:G2')
    sheet.merge_cells('H2:J2')

    sheet['A2'] = f"车牌号：{vehicle.license}"
    sheet['C2'] = f"驾驶员：{vehicle.driver}"
    sheet['D2'] = f"电话：{vehicle.phone}"
    sheet['E2'] = f"起始日期：{start_date}"
    sheet['H2'] = f"截止日期：{end_date}"

    for row in sheet['A2:J2']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = thin_border
    headers = ["序号", "日期", "运输起点", "运输终点", "品类", "数量", "单位", "装车方式", "给司机单价", "金额"]
    sheet.append(headers)
    current_row = sheet.max_row
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    total_amount = 0.0
    for idx, item in enumerate(items, start=1):
        start_site = Site.objects.filter(id=item.startsite_id).first()
        start_site_name = start_site.name if start_site else "无"
        end_site = Site.objects.filter(id=item.endsite_id).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        total_price = item.quantity * item.driverPrice
        row = [
            idx,
            item.date.split('T')[0],
            start_site_name,
            end_site_name,
            goods_name,
            item.quantity,
            item.unit,
            item.get_load_display(),
            item.driverPrice,
            total_price
        ]
        sheet.append(row)
        current_row = sheet.max_row
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        total_amount += total_price
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append([f"总 计 金 额：{total_amount}", "", "", "", f"总 计 大 写 (金 额)：{total_cn}"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:D{current_row}')
    sheet.merge_cells(f'E{current_row}:J{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.border = thin_border

    sheet.row_dimensions[1].height = 25
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 18.5

    sheet.page_setup.scale = 69

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
        workbook.save(tmp_xlsx.name)
        tmp_xlsx_path = tmp_xlsx.name

    pdf_path = tmp_xlsx_path.replace('.xlsx', '.pdf')

    subprocess.run(['unoconv', '-f', 'pdf', tmp_xlsx_path])

    with open(pdf_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename=transport_statement.pdf'

    os.remove(tmp_xlsx_path)
    os.remove(pdf_path)

    return response




@CheckRequire
def payment(req:HttpRequest):
    # failure_response, user = get_user_from_request(req,'POST')
    # if failure_response:
    #     return failure_response
    body = json.loads(req.body.decode("utf-8"))
    owner = require(body,"owner","string",err_msg="Missing or error type of [owner]")
    date = require(body,"date","string",err_msg="Missing or error type of [date]")
    amount = require(body,"amount","float",err_msg="Missing or error type of [amount]")
    pay_id = require(body,"pay_id","int",err_msg="Missing or error type of [pay_id]")
    balance_amount = require(body,"balance_amount","float",err_msg="Missing or error type of [balance_amount]")
    try:
        note = require(body,"note","string",err_msg="Missing or error type of [note]")
    except:
        note = "无"
    Newpayment = Payment.objects.create(owner=owner,date=date,amount=amount,pay_id=pay_id,balance_amount=balance_amount,created_time=get_timestamp(),note=note,if_delete=False)
    return request_success()

@CheckRequire
def del_payment(req:HttpRequest,payment_id):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    payment = Payment.objects.filter(id=payment_id,if_delete=False).first()
    if not payment:
        return request_failed(code=1,info="Payment does not exist",status_code=404)
    payment.if_delete=True
    payment.save()
    return request_success()

@CheckRequire
def payment_list(req:HttpRequest,per_page,page):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    ownerName = req.GET.get('ownerName',None)
    start_date = req.GET.get('start_date',None)
    end_date = req.GET.get('end_date',None)
    payments = Payment.objects.filter(if_delete=False)
    if ownerName:
        payments = payments.filter(owner=ownerName)
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)
        
    paginator = Paginator(payments, per_page)
    current_page = paginator.get_page(page)
    total_pages = paginator.num_pages
    return_data = [payment.serialize() for payment in current_page]
    return request_success({"payments":return_data,"total_pages":total_pages})

@CheckRequire
def change_payment(req:HttpRequest):
    # failure_response, user = get_user_from_request(req,'DELETE')
    # if failure_response:
    #     return failure_response
    body = json.loads(req.body.decode("utf-8"))
    payment_id = require(body,"payment_id","int",err_msg="Missing or error type of [payment_id]")
    payment = Payment.objects.filter(id=payment_id).first()
    if not payment:
        return request_failed(code=1,info="Payment does not exist",status_code=404)
    try:
        owner = require(body, "owner", "string", err_msg="Missing or error type of [owner]")
    except:
        owner = None
    try:
        pay_id = require(body, "pay_id", "int", err_msg="Missing or error type of [pay_id]")
    except:
        pay_id = None
    try:
        amount = require(body, "amount", "int", err_msg="Missing or error type of [amount]")
    except:
        amount = None
    try:
        date = require(body, "date", "string", err_msg="Missing or error type of [date]")
    except:
        date = None
    try:
        note = require(body, "note", "string", err_msg="Missing or error type of [note]")
    except:
        note = None
    try:
        balance_amount = require(body, "balance_amount", "float", err_msg="Missing or error type of [balance_amount]")
    except:
        balance_amount = None
    if owner:
        payment.owner = owner
    if amount:
        payment.amount = amount
    if date:
        payment.date = date
    if note:
        payment.note = note
    if pay_id:
        payment.pay_id = pay_id
    if balance_amount:
        payment.balance_amount = balance_amount
    payment.save()
    return request_success()

@CheckRequire
def search4payment(req:HttpRequest,per_page,page):
    start_date = req.GET.get('start_date',None)
    end_date = req.GET.get('end_date',None)
    owner = req.GET.get('owner',None)
    payments = Payment.objects.filter(if_delete=False).order_by("-created_time")
    if owner:
        payments = payments.filter(owner=owner)
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)
        
    paginator = Paginator(payments, per_page)
    current_page = paginator.get_page(page)
    total_pages = paginator.num_pages
    return_data = [payment.serialize() for payment in current_page]

    return request_success({"payments":return_data,"total_pages":total_pages})

