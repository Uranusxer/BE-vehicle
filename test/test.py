from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Sample data
startsite = {"id": 1, "name": "起始工地A", "owner": "老板A"}
endsites = {2: {"id": 2, "name": "终点工地B"}}
goods = {1: {"id": 1, "name": "土方"}, 2: {"id": 2, "name": "沙石"}}
items = [
    {"id": 1, "start_date": "2024-07-01T00:00:00Z", "start_spot": "A起点", "unit": "吨", "contractorPrice": 100, "goods_id": 1, "endsite_id": 2, "quantity": 5, "endPayment": 500, "startSubsidy": 50, "endSubsidy": 30},
    {"id": 2, "start_date": "2024-07-02T00:00:00Z", "start_spot": "A起点", "unit": "吨", "contractorPrice": 100, "goods_id": 1, "endsite_id": 2, "quantity": 10, "endPayment": 1000, "startSubsidy": 100, "endSubsidy": 60},
    {"id": 3, "start_date": "2024-07-03T00:00:00Z", "start_spot": "A起点", "unit": "吨", "contractorPrice": 120, "goods_id": 2, "endsite_id": 2, "quantity": 8, "endPayment": 960, "startSubsidy": 80, "endSubsidy": 40}
]
start_date = "2024-07-01T00:00:00Z"
end_date = "2024-07-02T00:00:00Z"
# 创建Excel工作簿
workbook = Workbook()
sheet = workbook.active
sheet.title = "宏途运输每月对账单"

# 设置列宽
column_widths = {
    'A': 5,    # 序号
    'B': 15,   # 起始日期
    'C': 15,   # 合并单元格"起始日期"
    'D': 20,   # 运输起点
    'E': 20,   # 合并单元格"运输起点"
    'F': 10,   # 终点工地
    'G': 15,   # 品类
    'H': 15,   # 合并单元格"品类"
    'I': 10,   # 数量
    'J': 10,   # 单位
    'K': 10,   # 单价
    'L': 15,   # 终点付费金额
    'M': 15,   # 起点补贴金额
    'N': 15    # 终点补贴金额
}

for col_letter, width in column_widths.items():
    sheet.column_dimensions[col_letter].width = width

# 添加表头
sheet.merge_cells('A1:N1')
title_cell = sheet['A1']
title_cell.value = "宏途运输每月对账单"
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.font = Font(size=24, bold=True)

# 空行


# 固定的表头信息
sheet.merge_cells('A2:C2')
sheet.merge_cells('D2:F2')
sheet.merge_cells('G2:H2')
sheet.merge_cells('I2:N2')
sheet.merge_cells('A3:C3')
sheet.merge_cells('D3:F3')
sheet.merge_cells('G3:H3')
sheet.merge_cells('I3:N3')
sheet.merge_cells('A4:C4')
sheet.merge_cells('D4:F4')
sheet.merge_cells('G4:H4')
sheet.merge_cells('I4:N4')

sheet['A2'] = "起 点 工 地 单 位 名 称"
sheet['D2'] = startsite['name']
sheet['G2'] = "工 地 老 板 名 称"
sheet['I2'] = startsite['owner']
sheet['A3'] = "对 账 起 始 日 期"
sheet['D3'] = start_date
sheet['G3'] = "对 账 截 止 日 期"
sheet['I3'] = end_date
sheet['A4'] = "运 输 单 位 名 称"
sheet['D4'] = "八 达 通 渣 土 运 输 有 限 公 司"
sheet['G4'] = "公 司 负 责 人"
sheet['I4'] = "叶 家 荣 19859999999"

for row in sheet['A2:N2']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
for cell in sheet['A3:N3']:
    cell.alignment = Alignment(horizontal='center', vertical='center')
for cell in sheet['A4:N4']:
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 列标题
headers = ["序号", "起始日期", "", "运输起点", "", "终点工地", "品类", "", "数量", "单位", "单价", "终点付费金额", "起点补贴金额", "终点补贴金额"]
sheet.append(headers)

current_row = sheet.max_row
sheet.merge_cells(f'B{current_row}:C{current_row}')
sheet.merge_cells(f'D{current_row}:E{current_row}')
sheet.merge_cells(f'G{current_row}:H{current_row}')

for cell in sheet[current_row]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 合并相同条件下的明细
summary = items.values(
    'start_date',
    'start_spot',
    'unit',
    'contractorPrice',
    'goods_id',
    'endsite_id'
).annotate(
    quantity_sum=Sum('quantity'),
    end_payment_sum=Sum('endPayment'),
    start_subsidy_sum=Sum('startSubsidy'),
    end_subsidy_sum=Sum('endSubsidy')
)

total_amount = 0
for idx, item in enumerate(summary, start=1):
    row = [
        idx,
        item['start_date'],
        "",
        item['start_spot'],
        "",
        Site.objects.filter(id=item['endsite_id']).first().name,
        Goods.objects.filter(id=item['goods_id']).first().name,
        "",
        item['quantity_sum'],
        item['unit'],
        item['contractorPrice'],
        item['end_payment_sum'],
        item['start_subsidy_sum'],
        item['end_subsidy_sum']
    ]
    # 将数据追加到sheet中，并合并相应的单元格
    sheet.append(row)
    current_row = sheet.max_row
    sheet.merge_cells(f'B{current_row}:C{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    total_amount += item['quantity_sum'] * item['contractorPrice']

# 合计行
total_cn = num2cn(total_amount)
sheet.append(["合 计", "", "", "-", "", "-", "", "-", sum(item['quantity_sum'] for item in summary), "-", "-", "-", "-", "-"])
current_row = sheet.max_row
sheet.merge_cells(f'A{current_row}:C{current_row}')
sheet.merge_cells(f'D{current_row}:E{current_row}')
sheet.merge_cells(f'G{current_row}:H{current_row}')
sheet.append(["总 计 金 额", "", "", total_amount, "", "总 计 大 写  ( 金 额 ）", "", "", total_cn])
current_row = sheet.max_row
sheet.merge_cells(f'A{current_row}:C{current_row}')
sheet.merge_cells(f'D{current_row}:E{current_row}')
sheet.merge_cells(f'F{current_row}:H{current_row}')
sheet.merge_cells(f'I{current_row}:N{current_row}')



# 运输品类合计
headers = ["","","序号", "运输起点", "", "终点工地", "品类", "", "数量", "单位", "单价", "终点付费金额", "起点补贴金额", "终点补贴金额"]
sheet.append(headers)

current_row = sheet.max_row
sheet.merge_cells(f'A{current_row}:B{current_row}')
sheet.merge_cells(f'D{current_row}:E{current_row}')
sheet.merge_cells(f'G{current_row}:H{current_row}')

for cell in sheet[current_row]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

transport_summary = items.values(
    'start_spot',
    'contractorPrice',
    'unit',
    'goods_id',
    'endsite_id'
).annotate(
    quantity_sum=Sum('quantity'),
    end_payment_sum=Sum('endPayment'),
    start_subsidy_sum=Sum('startSubsidy'),
    end_subsidy_sum=Sum('endSubsidy')
)
row1 = sheet.max_row+1
for idx, item in enumerate(transport_summary, start=1):
    row = [
        "",
        "",
        idx,
        item['start_spot'],
        "",
        Site.objects.filter(id=item['endsite_id']).first().name,
        Goods.objects.filter(id=item['goods_id']).first().name,
        "",
        item['quantity_sum'],
        item['unit'],
        item['contractorPrice'],
        item['end_payment_sum'],
        item['start_subsidy_sum'],
        item['end_subsidy_sum']
    ]
    sheet.append(row)
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:B{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
row2 = sheet.max_row
sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=2)
cell = sheet.cell(row=row1, column=1)
cell.value = "运 输 品 类 合 计"
cell.alignment = Alignment(horizontal='center', vertical='center')

sheet.append(["","","合计","-","","-","-","","-","-","-","-","-","-"])
current_row = sheet.max_row
sheet.merge_cells(f'A{current_row}:B{current_row}')
sheet.merge_cells(f'D{current_row}:E{current_row}')
sheet.merge_cells(f'G{current_row}:H{current_row}')

# 工地负责人及固定信息
sheet.append([""] * 11)
sheet.append(["工 地 负 责 人（ 签 字 确 认 ) ：","","","","","运 输 单 位 负 责 人 (  签 字 确 认 ) ："])
current_row = sheet.max_row
sheet.merge_cells(f'A{current_row}:E{current_row}')
sheet.merge_cells(f'F{current_row}:N{current_row}')


sheet.append(["经 营 范 围 ： 本 公 司 承 拆 土 石 方 工 程 、渣 土 、 建 筑 垃 圾 运 输 、 砂 石 料 、 柴 油 配 送 等 。"])
sheet.append(["成 就 伙 伴 、 铸 造 品 牌 、 我 们 每 天 努 力 、 我 们 每 天 进 步 。宏 途 运 输 每 月 对 帐 单"])
centered_row = sheet.max_row
# 设置该行每个单元格居中对齐
for cell in sheet[centered_row]:
    cell.alignment = Alignment(horizontal='center', vertical='center')
# Save workbook to file
local_file_path = "transport_statement.xlsx"
workbook.save(local_file_path)
print(f"文件已保存到本地: {local_file_path}")