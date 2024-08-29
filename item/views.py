from django.shortcuts import render
from item.models import Item
from parameter.models import Site,Goods,Vehicle,Project
from utils.utils_request import BAD_METHOD, request_failed, request_success
from utils.utils_require import CheckRequire, require
from utils.utils_time import get_timestamp
from django.core.paginator import Paginator
from user.models import User,get_user_from_request
from utils.constants import START,END
import json
from django.http import JsonResponse
from openpyxl import Workbook
import io
from datetime import datetime
from django.http import HttpRequest, HttpResponse
from openpyxl.styles import Alignment, Font, Border, Side
import re
from django.db.models import Sum, F, FloatField, Value
from django.contrib.postgres.aggregates import ArrayAgg
from django.db.models.functions import Coalesce
from collections import defaultdict
from datetime import datetime
import pytz
import tempfile
import pypandoc
import os
import subprocess
import fitz
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
import cn2an

def convert_utc_to_china_time(utc_time_str):
    # 将字符串转换为 datetime 对象
    utc_time = datetime.strptime(utc_time_str, '%Y-%m-%dT%H:%M:%S.%fZ')
    
    # 设置时区为 UTC
    utc_time = utc_time.replace(tzinfo=pytz.utc)
    
    # 转换到东八区中国时间
    china_time = utc_time.astimezone(pytz.timezone('Asia/Shanghai'))
    
    # 仅保留日期部分
    china_date_str = china_time.strftime('%Y-%m-%d')
    
    return china_date_str

@CheckRequire
def transport_item(req:HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    body = json.loads(req.body.decode("utf-8"))

    # all kinds of ids
    startsite_id = require(body,"startsite_id","int",err_msg="Missing or error type of [startsite_id]")
    endsite_id = require(body,"endsite_id","int",err_msg="Missing or error type of [endsite_id]")
    vehicle_id  = require(body,"vehicle_id","int",err_msg="Missing or error type of [vehicle_id]")
    goods_id = require(body,"goods_id","int",err_msg="Missing or error type of [goods_id]")
    project_id = require(body,"project_id","int",err_msg="Missing or error type of [project_id]")
    
    try:
        note = require(body,"note","string",err_msg="Missing or error type of [note]")
    except:
        note = None
    date = require(body,"date","string",err_msg="Missing or error type of [date]")
    unit = require(body,"unit","string",err_msg="Missing or error type of [unit]")
    quantity = require(body,"quantity","float",err_msg="Missing or error type of [quantity]")
    load = require(body,"load","string",err_msg="Missing or error type of [load]")

    # price
    try:
        contractorPrice = require(body, "contractorPrice", "float", err_msg="Missing or error type of [contractorPrice]")
    except:
        contractorPrice = 0
    try:
        startSubsidy = require(body, "startSubsidy", "float", err_msg="Missing or error type of [startSubsidy]")
    except:
        startSubsidy = 0
    try:
        endSubsidy = require(body, "endSubsidy", "float", err_msg="Missing or error type of [endSubsidy]")
    except:
        endSubsidy = 0
    try:
        endPayment = require(body, "endPayment", "float", err_msg="Missing or error type of [endPayment]")
    except:
        endPayment = 0
    try:
        driverPrice = require(body, "driverPrice", "float", err_msg="Missing or error type of [driverPrice]")
    except:
        driverPrice = 0

    Newitem = Item.objects.create(startsite_id=startsite_id,endsite_id=endsite_id,vehicle_id=vehicle_id,
            goods_id=goods_id,project_id=project_id,date=date,unit=unit,quantity=quantity,note=note,
            load=load,contractorPrice=contractorPrice,startSubsidy=startSubsidy,endSubsidy=endSubsidy,
            endPayment=endPayment,driverPrice=driverPrice,created_time=get_timestamp())
    return request_success()

@CheckRequire
def del_item(req:HttpRequest,item_id):
    failure_response, user = get_user_from_request(req,'DELETE')
    if failure_response:
        return failure_response
    item = Item.objects.filter(id=item_id,if_delete=False).first()
    if not item:
        request_failed(code=1,info="Item does not exist",status_code=404)
    item.if_delete=True
    item.save()
    return request_success()

@CheckRequire
def change_item(req:HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response

    body = json.loads(req.body.decode("utf-8"))
    item_id = require(body,"item_id","int",err_msg="Missing or error type of [item_id]")
    item = Item.objects.filter(id=item_id,if_delete=False).first()
    if not item:
        return request_failed(code=1,info="Item does not exist",status_code=404)
    try:
        startsite_id = require(body, "startsite_id", "int", err_msg="Missing or error type of [startsite_id]")
    except:
        startsite_id = None
    try:
        endsite_id = require(body, "endsite_id", "int", err_msg="Missing or error type of [endsite_id]")
    except:
        endsite_id = None
    try:
        vehicle_id = require(body, "vehicle_id", "int", err_msg="Missing or error type of [vehicle_id]")
    except:
        vehicle_id = None
    try:
        goods_id = require(body, "goods_id", "int", err_msg="Missing or error type of [goods_id]")
    except:
        goods_id = None
    try:
        project_id = require(body, "project_id", "int", err_msg="Missing or error type of [project_id]")
    except:
        project_id = None

    try:
        date = require(body, "date", "string", err_msg="Missing or error type of [date]")
    except:
        date = None
    try:
        unit = require(body, "unit", "string", err_msg="Missing or error type of [unit]")
    except:
        unit = None
    try:
        quantity = require(body, "quantity", "float", err_msg="Missing or error type of [quantity]")
    except:
        quantity = None
    try:
        note = require(body, "note", "string", err_msg="Missing or error type of [note]")
    except:
        note = None
    try:
        load = require(body, "load", "string", err_msg="Missing or error type of [load]")
    except:
        load = None

    try:
        contractorPrice = require(body, "contractorPrice", "float", err_msg="Missing or error type of [contractorPrice]")
    except:
        contractorPrice = None
    try:
        startSubsidy = require(body, "startSubsidy", "float", err_msg="Missing or error type of [startSubsidy]")
    except:
        startSubsidy = None
    try:
        endSubsidy = require(body, "endSubsidy", "float", err_msg="Missing or error type of [endSubsidy]")
    except:
        endSubsidy = None
    try:
        endPayment = require(body, "endPayment", "float", err_msg="Missing or error type of [endPayment]")
    except:
        endPayment = None
    try:
        driverPrice = require(body, "driverPrice", "float", err_msg="Missing or error type of [driverPrice]")
    except:
        driverPrice = None


    if startsite_id:
        item.startsite_id = startsite_id
    if endsite_id:
        item.endsite_id = endsite_id
    if vehicle_id:
        item.vehicle_id = vehicle_id
    if goods_id:
        item.goods_id = goods_id
    if project_id:
        item.project_id = project_id

    if date:
        item.date = date
    if unit:
        item.unit = unit
    if quantity:
        item.quantity = quantity
    if note:
        item.note = note
    if load:
        item.load = load

    if contractorPrice:
        item.contractorPrice = contractorPrice
    if startSubsidy:
        item.startSubsidy = startSubsidy
    if endSubsidy:
        item.endSubsidy = endSubsidy
    if endPayment:
        item.endPayment = endPayment
    if driverPrice:
        item.driverPrice = driverPrice

    item.save()
    return request_success()

@CheckRequire
def search4item(req:HttpRequest,per_page,page):
    failure_response, user = get_user_from_request(req,'GET')
    if failure_response:
        return failure_response
    project_owner = req.GET.get('ownerName', None)
    startsite_id = req.GET.get('startsite_id',None)
    endsite_id = req.GET.get('endsite_id',None)
    vehicle_id = req.GET.get('vehicle_id',None)
    goods_id = req.GET.get('goods_id',None)
    project_id = req.GET.get('project_id', None)
    start_date = req.GET.get('start_date',None)
    end_date = req.GET.get('end_date',None)


    items = Item.objects.filter(if_delete=False)
    if project_owner is not None: 
        project_ids = Project.objects.filter(owner=project_owner).values_list('id', flat=True)
        items = items.filter(project_id__in=project_ids)
    
    if startsite_id is not None:
        items = items.filter(startsite_id=startsite_id)
    if endsite_id is not None:
        items = items.filter(endsite_id=endsite_id)
    if vehicle_id is not None:
        items = items.filter(vehicle_id=vehicle_id)

    if goods_id is not None:
        items = items.filter(goods_id=goods_id)
    if project_id is not None:
        items = items.filter(project_id=project_id)

    if start_date is not None:
        items = items.filter(date__gte=start_date)
    if end_date is not None:
        items = items.filter(date__lte=end_date)

        
    paginator = Paginator(items, per_page)
    current_page = paginator.get_page(page)
    total_pages = paginator.num_pages
    return_data = [item.serialize() for item in current_page]

    return request_success({"items":return_data,"total_pages":total_pages})

@CheckRequire
def item_list(req:HttpRequest,per_page,page):
    failure_response, user = get_user_from_request(req,'GET')
    if failure_response:
        return failure_response
    item_list = Item.objects.filter(if_delete=False).order_by("-created_time")
    paginator = Paginator(item_list,per_page)
    item_page = paginator.get_page(page)
    return_data = [item.serialize() for item in item_page]
    total_pages = paginator.num_pages
    return request_success({"items":return_data,"total_pages":total_pages})



@CheckRequire
def item_price(req: HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    body = json.loads(req.body.decode("utf-8"))

    items = body.get('items', [])
    if not isinstance(items, list):
        return JsonResponse({"error": "Invalid data format, 'items' should be a list"}, status=400)

    response_data = {"updated_items": [], "errors": []}

    for item_data in items:
        try:
            item_id = require(item_data, "item_id", "int", err_msg="Missing or error type of [item_id]")
            try:
                contractorPrice = require(item_data, "contractorPrice", "float", err_msg="Missing or error type of [contractorPrice]")
            except:
                contractorPrice = None
            try:
                startSubsidy = require(item_data, "startSubsidy", "float", err_msg="Missing or error type of [startSubsidy]")
            except:
                startSubsidy = None
            try:
                endSubsidy = require(item_data, "endSubsidy", "float", err_msg="Missing or error type of [endSubsidy]")
            except:
                endSubsidy = None
            try:
                endPayment = require(item_data, "endPayment", "float", err_msg="Missing or error type of [endPayment]")
            except:
                endPayment = None
            try:
                driverPrice = require(item_data, "driverPrice", "float", err_msg="Missing or error type of [driverPrice]")
            except:
                driverPrice = None
            try:
                quantity = require(item_data, "quantity", "float", err_msg="Missing or error type of [quantity]")
            except:
                quantity = None
            try:
                unit = require(item_data, "unit", "string", err_msg="Missing or error type of [unit]")
            except:
                unit = None

            item = Item.objects.filter(id=item_id,if_delete=False).first()
            if not item:
                response_data["errors"].append({"item_id": item_id, "error": "Item does not exist"})
                continue
            if contractorPrice or contractorPrice == 0:
                item.contractorPrice = contractorPrice
            if startSubsidy or startSubsidy == 0:
                item.startSubsidy = startSubsidy
            if endSubsidy or endSubsidy == 0:
                item.endSubsidy = endSubsidy
            if endPayment or endPayment == 0:
                item.endPayment = endPayment
            if driverPrice or driverPrice == 0:
                item.driverPrice = driverPrice
            if quantity or quantity == 0:
                item.quantity = quantity
            if unit:
                item.unit = unit
            item.save()

            response_data["updated_items"].append(item_id)
        
        except Exception as e:
            response_data["errors"].append({
                "item_data": item_data,
                "error": str(e)
            })
    
    if response_data["errors"]:
        return request_success()
    else:
        return JsonResponse(response_data, status=200)




@CheckRequire
def start_excel(req: HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    # 从请求参数中获取数据
    body = json.loads(req.body.decode("utf-8"))
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    project_id = require(body, "project_id", "int", err_msg="Missing or error type of [project_id]")
    start_date = require(body, "start_date", "string", err_msg="Missing or error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing or error type of [end_date]")
    
    # 获取startsite对象并检查
    if project_id == 0:
        project = None
    else:   
        project = Project.objects.filter(id=project_id).first()
        if not project:
            return request_failed(code=3, info="Project not found", status_code=404)
    
    # 转换日期格式，假设传入的日期格式为"2024-07-03T16:00:00.000Z"
    start_date = convert_utc_to_china_time(start_date).split('T')[0]
    end_date = convert_utc_to_china_time(end_date).split('T')[0]
    
    # 获取过滤的Item
    items = Item.objects.filter(id__in=item_ids, if_delete=False)
    
    # 创建Excel工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运公司对账单"
    default_font = Font(size=12)
    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置列宽
    column_widths = {
        'A': 10.41,    # 序号
        'B': 8.17,   # 起始日期
        'C': 11.29,   # 合并单元格"起始日期"
        'D': 26.29,   # 运输起点
        'E': 12.14,   # 品类
        'F': 40.15,   # 项目老板名
        'G': 12.68,   # 装车方式
        'H': 8.49,   # 数量
        'I': 8.49,   # 单位
        'J': 16.87,   # 工地承接单价
        'K': 10.58,   # 总金额
        'L': 16.87,   # 起点补贴金额
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # 添加表头
    sheet.merge_cells('A1:L1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运公司对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)


    # 固定的表头信息
    sheet.merge_cells('A2:C2')
    sheet.merge_cells('D2:E2')
    sheet.merge_cells('G2:L2')
    sheet.merge_cells('A3:C3')
    sheet.merge_cells('D3:E3')
    sheet.merge_cells('G3:L3')
    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D4:E4')
    sheet.merge_cells('G4:L4')

    
    sheet['A2'] = "项 目 名 称"
    sheet['D2'] = project.name if project else None
    sheet['F2'] = "项 目 老 板 名 称"
    sheet['G2'] = project.owner if project else None
    sheet['A3'] = "对 账 起 始 日 期"
    sheet['D3'] = start_date
    sheet['F3'] = "对 账 截 止 日 期"
    sheet['G3'] = end_date
    sheet['A4'] = "运 输 单 位 名 称"
    sheet['D4'] = "南平市宏途渣土清运有限公司"
    sheet['F4'] = "公 司 负 责 人"
    sheet['G4'] = "吴 春 才 18905996295"
    
    for row in sheet['A2:L4']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
    
    # 列标题
    headers = ["序号", "日期", "", "运输起点", "品类", "车队", "装车方式", "数量", "单位", "工地承接单价", "总金额", "起点补贴金额"]
    sheet.append(headers)
    
    current_row = sheet.max_row
    sheet.merge_cells(f'B{current_row}:C{current_row}')
    
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 不合并
    
    total_amount = 0
    for idx, item in enumerate(items, start=1):
        # get necessary info
        start_site = Site.objects.filter(id=item.startsite_id).first()
        start_site_name = start_site.name if start_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        vehicle = Vehicle.objects.filter(id=item.vehicle_id).first()
        total_price = item.quantity * item.contractorPrice
        row = [
            idx,
            convert_utc_to_china_time(item.date).split('T')[0],
            "",
            start_site_name,
            goods_name,
            vehicle.license,
            item.get_load_display(),
            item.quantity,
            item.unit,
            item.contractorPrice,
            total_price,
            item.startSubsidy
        ]
        # 将数据追加到sheet中，并合并相应的单元格
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'B{current_row}:C{current_row}')

        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        total_amount += total_price

    # 合计行
    # sheet.append(["合 计", "", "", "-", "", "-", "", "-", sum(item['quantity_sum'] for item in summary), "-", "-", "-", "-", "-"])
    # current_row = sheet.max_row
    # sheet.merge_cells(f'A{current_row}:C{current_row}')
    # sheet.merge_cells(f'D{current_row}:E{current_row}')
    # sheet.merge_cells(f'G{current_row}:H{current_row}')
    # for cell in sheet[current_row]:
    #     cell.alignment = Alignment(h  orizontal='center')
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append(["总 计 金 额", "", "", total_amount, "", "总 计 大 写 (金 额)",total_cn])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:C{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
    # 运输品类合计
    headers = ["","","序号", "运输起点", "品类", "车队", "合计数量","","单位","工地承接单价","总金额","起点补贴金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 初步查询，获取所需的基本信息
    initial_query = items.values(
        'startsite_id',
        'goods_id',
        'startSubsidy',
        'quantity',
        'contractorPrice',
        'vehicle_id',
        'unit',
    )

    # 准备聚合结果的字典
    transport_summary_dict = defaultdict(lambda: {
        'start_subsidy_sum': 0,
        'cost_sum': 0,
        'quantity_sum': 0,
        'vehicle_ids': []
    })

    # 循环处理每个初步查询的结果
    for item in initial_query:
        key = (item['startsite_id'], item['goods_id'], item['contractorPrice'], item['unit'])
        transport_summary_dict[key]['start_subsidy_sum'] += item['startSubsidy']
        transport_summary_dict[key]['cost_sum'] += item['quantity'] * item['contractorPrice']
        transport_summary_dict[key]['quantity_sum'] += item['quantity']
        transport_summary_dict[key]['vehicle_ids'].append(item['vehicle_id'])

    # 将聚合结果转换为列表
    transport_summary = []
    for (startsite_id, goods_id, contractorPrice, unit), data in transport_summary_dict.items():
        transport_summary.append({
            'startsite_id': startsite_id,
            'goods_id': goods_id,
            'start_subsidy_sum': data['start_subsidy_sum'],
            'quantity_sum': data['quantity_sum'],
            'cost_sum': data['cost_sum'],
            'vehicle_ids': data['vehicle_ids'],
            "contractorPrice": contractorPrice,
            "unit": unit
        })


    row1 = sheet.max_row
    total_sum1 = 0.0
    total_sum2 = 0.0
    for idx, item in enumerate(transport_summary, start=1):
        start_site = Site.objects.filter(id=item['startsite_id']).first()
        start_site_name = start_site.name if start_site else "无"
        goods = Goods.objects.filter(id=item['goods_id']).first()
        goods_name = goods.name if goods else "无"
        vehicle_ids = item['vehicle_ids'][:3]  # Extract up to the first three vehicle IDs
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
        vehicle_names ='，'.join([vehicle.license for vehicle in vehicles])
        # Accumulate sums
        total_sum1 += item['cost_sum']
        total_sum2 += item['start_subsidy_sum']
        row = [
            "",
            "",
            idx,
            start_site_name,
            goods_name,
            vehicle_names,
            item['quantity_sum'],
            "",
            item['unit'],
            item['contractorPrice'],
            item['cost_sum'],
            item['start_subsidy_sum']
        ]
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'G{current_row}:H{current_row}')
        
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')


    sheet.append(["","","合计","-","-","-",total_sum1,"","","","",total_sum2])
    
    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center')

    # 运输品类合计单元格
    row2 = sheet.max_row
    sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=2)
    cell = sheet.cell(row=row1, column=1)
    cell.value = "运输品类合计"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 工地负责人及固定信息
    sheet.append(["工 地 负 责 人（ 签 字 确 认 ) ：","","","","","运 输 单 位 负 责 人 (  签 字 确 认 ) ："])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:E{current_row}')
    sheet.merge_cells(f'F{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    sheet.append(["经 营 范 围 ： 建筑垃圾清运，砂石料运输及销售，供应铺路石渣，云梯车租赁。"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
    sheet.append(["立 信 于 心 ， 尽 责 至 善！"])
    centered_row = sheet.max_row
    sheet.merge_cells(f'A{centered_row}:L{centered_row}')
    # 设置该行每个单元格居中对齐
    for cell in sheet[centered_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=16,bold=True)

    # # 保存到本地
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18.5
        for cell in row:
            if not cell.font:
                set_font_and_alignment(cell)
    sheet.row_dimensions[1].height = 25  
    sheet.row_dimensions[2].height = 25  
    sheet.row_dimensions[3].height = 25  
    sheet.row_dimensions[sheet.max_row].height = 22  
    sheet.row_dimensions[sheet.max_row-2].height = 30  
    # return request_success()

    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment;filename=transport_statement.xlsx'
    return response

@CheckRequire
def start_excel_pdf(req: HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    body = json.loads(req.body.decode("utf-8"))
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    project_id = require(body, "project_id", "int", err_msg="Missing or error type of [project_id]")
    start_date = require(body, "start_date", "string", err_msg="Missing或error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing或error type of [end_date]")

    if project_id == 0:
        project = None
    else:
        project = Project.objects.filter(id=project_id).first()
        if not project:
            return request_failed(code=3, info="Project not found", status_code=404)

    start_date = convert_utc_to_china_time(start_date).split('T')[0]
    end_date = convert_utc_to_china_time(end_date).split('T')[0]
    items = Item.objects.filter(id__in=item_ids, if_delete=False)
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运公司对账单"
    default_font = Font(size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    column_widths = {
        'A': 10.41,    # 序号
        'B': 8.17,   # 起始日期
        'C': 11.29,   # 合并单元格"起始日期"
        'D': 26.29,   # 运输起点
        'E': 12.14,   # 品类
        'F': 40.15,   # 项目老板名
        'G': 12.68,   # 装车方式
        'H': 8.49,   # 数量
        'I': 8.49,   # 单位
        'J': 16.87,   # 工地承接单价
        'K': 10.58,   # 总金额
        'L': 16.87,   # 起点补贴金额
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    sheet.merge_cells('A1:L1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运公司对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    title_cell.border = border

    # 添加表头
    sheet.merge_cells('A1:L1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运公司对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)


    # 固定的表头信息
    sheet.merge_cells('A2:C2')
    sheet.merge_cells('D2:E2')
    sheet.merge_cells('G2:L2')
    sheet.merge_cells('A3:C3')
    sheet.merge_cells('D3:E3')
    sheet.merge_cells('G3:L3')
    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D4:E4')
    sheet.merge_cells('G4:L4')

    
    sheet['A2'] = "项 目 名 称"
    sheet['D2'] = project.name if project else None
    sheet['F2'] = "项 目 老 板 名 称"
    sheet['G2'] = project.owner if project else None
    sheet['A3'] = "对 账 起 始 日 期"
    sheet['D3'] = start_date
    sheet['F3'] = "对 账 截 止 日 期"
    sheet['G3'] = end_date
    sheet['A4'] = "运 输 单 位 名 称"
    sheet['D4'] = "南平市宏途渣土清运有限公司"
    sheet['F4'] = "公 司 负 责 人"
    sheet['G4'] = "吴 春 才 18905996295"
    
    for row in sheet['A2:L4']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = border

    headers = ["序号", "日期", "", "运输起点", "品类", "车队", "装车方式", "数量", "单位", "工地承接单价", "总金额", "起点补贴金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'B{current_row}:C{current_row}')

    # Fill header row with font and alignment
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    total_amount = 0
    for idx, item in enumerate(items, start=1):
        start_site = Site.objects.filter(id=item.startsite_id).first()
        start_site_name = start_site.name if start_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        vehicle = Vehicle.objects.filter(id=item.vehicle_id).first()
        total_price = item.quantity * item.contractorPrice
        row = [
            idx,
            convert_utc_to_china_time(item.date).split('T')[0],
            "",
            start_site_name,
            goods_name,
            vehicle.license,
            item.get_load_display(),
            item.quantity,
            item.unit,
            item.contractorPrice,
            total_price,
            item.startSubsidy
        ]
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'B{current_row}:C{current_row}')

        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        total_amount += total_price
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append(["总 计 金 额", "", "", total_amount, "", "总 计 大 写 (金 额)", total_cn])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:C{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

   # 运输品类合计
    headers = ["","","序号", "运输起点", "品类", "车队", "合计数量","","单位","工地承接单价","总金额","起点补贴金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 初步查询，获取所需的基本信息
    initial_query = items.values(
        'startsite_id',
        'goods_id',
        'startSubsidy',
        'quantity',
        'contractorPrice',
        'vehicle_id',
        'unit',
    )

    # 准备聚合结果的字典
    transport_summary_dict = defaultdict(lambda: {
        'start_subsidy_sum': 0,
        'cost_sum': 0,
        'quantity_sum': 0,
        'vehicle_ids': []
    })

    # 循环处理每个初步查询的结果
    for item in initial_query:
        key = (item['startsite_id'], item['goods_id'], item['contractorPrice'], item['unit'])
        transport_summary_dict[key]['start_subsidy_sum'] += item['startSubsidy']
        transport_summary_dict[key]['cost_sum'] += item['quantity'] * item['contractorPrice']
        transport_summary_dict[key]['quantity_sum'] += item['quantity']
        transport_summary_dict[key]['vehicle_ids'].append(item['vehicle_id'])

    # 将聚合结果转换为列表
    transport_summary = []
    for (startsite_id, goods_id, contractorPrice, unit), data in transport_summary_dict.items():
        transport_summary.append({
            'startsite_id': startsite_id,
            'goods_id': goods_id,
            'start_subsidy_sum': data['start_subsidy_sum'],
            'quantity_sum': data['quantity_sum'],
            'cost_sum': data['cost_sum'],
            'vehicle_ids': data['vehicle_ids'],
            "contractorPrice": contractorPrice,
            "unit": unit
        })


    row1 = sheet.max_row
    total_sum1 = 0.0
    total_sum2 = 0.0
    for idx, item in enumerate(transport_summary, start=1):
        start_site = Site.objects.filter(id=item['startsite_id']).first()
        start_site_name = start_site.name if start_site else "无"
        goods = Goods.objects.filter(id=item['goods_id']).first()
        goods_name = goods.name if goods else "无"
        vehicle_ids = item['vehicle_ids'][:3]  # Extract up to the first three vehicle IDs
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
        vehicle_names ='，'.join([vehicle.license for vehicle in vehicles])
        # Accumulate sums
        total_sum1 += item['cost_sum']
        total_sum2 += item['start_subsidy_sum']
        row = [
            "",
            "",
            idx,
            start_site_name,
            goods_name,
            vehicle_names,
            item['quantity_sum'],
            "",
            item['unit'],
            item['contractorPrice'],
            item['cost_sum'],
            item['start_subsidy_sum']
        ]
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'G{current_row}:H{current_row}')
        
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    sheet.append(["","","合计","-","-","-",total_sum1,"","","","",total_sum2])
    
    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center')
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border


    row2 = sheet.max_row
    sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=2)
    cell = sheet.cell(row=row1, column=1)
    cell.value = "运输品类合计"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

    # Final Information rows
    sheet.append(["工 地 负 责 人（ 签 字 确 认 ) ：", "", "", "", "", "运 输 单 位 负 责 人 ( 签 字 确 认 ) ："])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:E{current_row}')
    sheet.merge_cells(f'F{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    sheet.append(["经 营 范 围 ： 建筑垃圾清运，砂石料运输及销售，供应铺路石渣，云梯车租赁。"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:L{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.border = border

    sheet.append(["立 信 于 心 ， 尽 责 至 善！"])
    centered_row = sheet.max_row
    sheet.merge_cells(f'A{centered_row}:L{centered_row}')
    for cell in sheet[centered_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=16, bold=True)
        cell.border = border

    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18.5
        for cell in row:
            if not cell.font:
                set_font_and_alignment(cell)
    sheet.row_dimensions[1].height = 25  
    sheet.row_dimensions[2].height = 25  
    sheet.row_dimensions[3].height = 25  
    sheet.row_dimensions[sheet.max_row].height = 22  
    sheet.row_dimensions[sheet.max_row-2].height = 30  

     # 增加缩放比例
    sheet.page_setup.scale = 48  # 设置缩放比例为 48%

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
        workbook.save(tmp_xlsx.name)
        tmp_xlsx_path = tmp_xlsx.name
    
    pdf_path = tmp_xlsx_path.replace('.xlsx', '.pdf')  # 临时 PDF 文件路径

    # 使用 unoconv 将 xlsx 转换为 pdf
    subprocess.run(['unoconv', '-f', 'pdf', tmp_xlsx_path])

    with open(pdf_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename=transport_statement.pdf'

    # 删除临时文件
    os.remove(tmp_xlsx_path)
    os.remove(pdf_path)  # 可选，保留 pdf 文件进行排查

    return response


@CheckRequire
def end_excel(req: HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    # 从请求参数中获取数据
    body = json.loads(req.body.decode("utf-8"))
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    project_id = require(body, "project_id", "int", err_msg="Missing or error type of [project_id]")
    start_date = require(body, "start_date", "string", err_msg="Missing or error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing or error type of [end_date]")
    
    # 获取startsite对象并检查
    if project_id == 0:
        project = None
    else:
        project = Project.objects.filter(id=project_id).first()
        if not project:
            return request_failed(code=3, info="Project not found", status_code=404)
    
    # 转换日期格式，假设传入的日期格式为"2024-07-03T16:00:00.000Z"
    start_date = convert_utc_to_china_time(start_date).split('T')[0]
    end_date = convert_utc_to_china_time(end_date).split('T')[0]
    
    # 获取过滤的Item
    items = Item.objects.filter(id__in=item_ids, if_delete=False)
    
    # 创建Excel工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运公司对账单"
    default_font = Font(size=12)
    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置列宽
    column_widths = {
        'A': 10.41,    # 序号
        'B': 8.17,   # 起始日期
        'C': 11.29,   # 合并单元格"起始日期"
        'D': 26.29,   # 运输起点
        'E': 12.14,   # 品类
        'F': 40.15,   # 项目老板名
        'G': 12.68,   # 装车方式
        'H': 8.49,   # 数量
        'I': 8.49,   # 单位
        'J': 16.87,   # 工地承接单价
        'K': 10.58,   # 总金额
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # 添加表头
    sheet.merge_cells('A1:K1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运公司对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)


    # 固定的表头信息
    sheet.merge_cells('A2:C2')
    sheet.merge_cells('D2:E2')
    sheet.merge_cells('G2:K2')
    sheet.merge_cells('A3:C3')
    sheet.merge_cells('D3:E3')
    sheet.merge_cells('G3:K3')
    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D4:E4')
    sheet.merge_cells('G4:K4')

    
    sheet['A2'] = "对 账 起 始 日 期"
    sheet['D2'] = start_date
    sheet['F2'] = "对 账 截 止 日 期"
    sheet['G2'] = end_date
    sheet['A3'] = "运 输 单 位 名 称"
    sheet['D3'] = "南平市宏途渣土清运有限公司"
    sheet['F3'] = "公 司 负 责 人"
    sheet['G3'] = "吴 春 才 18905996295"
    
    for row in sheet['A2:K3']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
    
    # 列标题
    headers = ["序号", "日期", "", "运输终点", "品类", "车队", "装车方式", "数量", "单位", "终点付费金额", "总金额"]
    sheet.append(headers)
    
    current_row = sheet.max_row
    sheet.merge_cells(f'B{current_row}:C{current_row}')
    
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 不合并
    
    total_amount = 0
    for idx, item in enumerate(items, start=1):
        # get necessary info
        end_site = Site.objects.filter(id=item.endsite_id).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        vehicle = Vehicle.objects.filter(id=item.vehicle_id).first()
        total_price = item.quantity * item.endPayment
        row = [
            idx,
            convert_utc_to_china_time(item.date).split('T')[0],
            "",
            end_site_name,
            goods_name,
            vehicle.license,
            item.get_load_display(),
            item.quantity,
            item.unit,
            item.endPayment,
            total_price
            ]
        # 将数据追加到sheet中，并合并相应的单元格
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'B{current_row}:C{current_row}')

        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        total_amount += total_price

    # 合计行
    # sheet.append(["合 计", "", "", "-", "", "-", "", "-", sum(item['quantity_sum'] for item in summary), "-", "-", "-", "-", "-"])
    # current_row = sheet.max_row
    # sheet.merge_cells(f'A{current_row}:C{current_row}')
    # sheet.merge_cells(f'D{current_row}:E{current_row}')
    # sheet.merge_cells(f'G{current_row}:H{current_row}')
    # for cell in sheet[current_row]:
    #     cell.alignment = Alignment(h  orizontal='center')
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append(["总 计 金 额", "", "", total_amount, "", "总 计 大 写 (金 额)",total_cn])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:C{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center')
    
    
    # 运输品类合计
    headers = ["","","序号", "运输终点", "品类", "车队", "合计数量","","单位","终点付费金额","总金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 初步查询，获取所需的基本信息
    initial_query = items.values(
        'endsite_id',
        'goods_id',
        'endPayment',
        'quantity',
        'vehicle_id',
        "unit"
    )

    # 准备聚合结果的字典
    transport_summary_dict = defaultdict(lambda: {
        'endPayment_sum': 0,
        'cost_sum': 0,
        'quantity_sum': 0,
        'vehicle_ids': []
    })

    # 循环处理每个初步查询的结果
    for item in initial_query:
        key = (item['endsite_id'], item['goods_id'], item['endPayment'], item['unit'])
        transport_summary_dict[key]['endPayment_sum'] += item['endPayment']
        transport_summary_dict[key]['cost_sum'] += item['quantity'] * item['endPayment']
        transport_summary_dict[key]['quantity_sum'] += item['quantity']
        transport_summary_dict[key]['vehicle_ids'].append(item['vehicle_id'])

    # 将聚合结果转换为列表
    transport_summary = []
    for (endsite_id, goods_id, endPayment, unit), data in transport_summary_dict.items():
        transport_summary.append({
            'endsite_id': endsite_id,
            'goods_id': goods_id,
            'endPayment_sum': data['endPayment_sum'],
            'quantity_sum': data['quantity_sum'],
            'cost_sum': data['cost_sum'],
            'vehicle_ids': data['vehicle_ids'],
            'endPayment': endPayment,
            'unit': unit
        })


    row1 = sheet.max_row    
    total_sum1 = 0.0
    total_sum2 = 0.0
    for idx, item in enumerate(transport_summary, start=1):
        end_site = Site.objects.filter(id=item['endsite_id']).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item['goods_id']).first()
        goods_name = goods.name if goods else "无"
        vehicle_ids = item['vehicle_ids'][:3]  # Extract up to the first three vehicle IDs
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
        vehicle_names ='，'.join([vehicle.license for vehicle in vehicles])
        # Accumulate sums
        total_sum1 += item['cost_sum']
        row = [
            "",
            "",
            idx,
            end_site_name,
            goods_name,
            vehicle_names,
            item['quantity_sum'],
            "",
            item['unit'],
            item['endPayment'],
            item['cost_sum'],
        ]
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'G{current_row}:H{current_row}')
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet.append(["","","合计","-","-","-",total_sum1,"","","",""])
    
    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center')

    # 运输品类合计单元格
    row2 = sheet.max_row
    sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=2)
    cell = sheet.cell(row=row1, column=1)
    cell.value = "运输品类合计"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 工地负责人及固定信息
    sheet.append(["工 地 负 责 人（ 签 字 确 认 ) ：","","","","","运 输 单 位 负 责 人 (  签 字 确 认 ) ："])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:E{current_row}')
    sheet.merge_cells(f'F{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    sheet.append(["经 营 范 围 ： 建筑垃圾清运，砂石料运输及销售，供应铺路石渣，云梯车租赁。"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
    sheet.append(["立 信 于 心 ， 尽 责 至 善！"])
    centered_row = sheet.max_row
    sheet.merge_cells(f'A{centered_row}:K{centered_row}')
    # 设置该行每个单元格居中对齐
    for cell in sheet[centered_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=16,bold=True)

    # # 保存到本地
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18.5
        for cell in row:
            if not cell.font:
                set_font_and_alignment(cell)
    sheet.row_dimensions[1].height = 25  
    sheet.row_dimensions[2].height = 25  
    sheet.row_dimensions[3].height = 25  
    sheet.row_dimensions[sheet.max_row].height = 22  
    sheet.row_dimensions[sheet.max_row-2].height = 30  
    # return request_success()

    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment;filename=transport_statement.xlsx'
    return response


@CheckRequire
def end_excel_pdf(req: HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    body = json.loads(req.body.decode("utf-8"))
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    project_id = require(body, "project_id", "int", err_msg="Missing or error type of [project_id]")
    start_date = require(body, "start_date", "string", err_msg="Missing or error type of [start_date]")
    end_date = require(body, "end_date", "string", err_msg="Missing or error type of [end_date]")

    if project_id == 0:
        project = None
    else:
        project = Project.objects.filter(id=project_id).first()
        if not project:
            return request_failed(code=3, info="Project not found", status_code=404)

    start_date = convert_utc_to_china_time(start_date).split('T')[0]
    end_date = convert_utc_to_china_time(end_date).split('T')[0]

    items = Item.objects.filter(id__in=item_ids, if_delete=False)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运公司对账单"

    default_font = Font(size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Function to set fonts, alignments, and borders
    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    column_widths = {
        'A': 10.41, 'B': 8.17, 'C': 11.29, 'D': 26.29, 'E': 12.14,
        'F': 40.15, 'G': 12.68, 'H': 8.49, 'I': 8.49, 'J': 16.87, 'K': 10.58
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    sheet.merge_cells('A1:K1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运公司对账单"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    title_cell.border = thin_border
    sheet['K1'].border = thin_border
    sheet.merge_cells('A2:C2')
    sheet.merge_cells('D2:E2')
    sheet.merge_cells('G2:K2')
    sheet.merge_cells('A3:C3')
    sheet.merge_cells('D3:E3')
    sheet.merge_cells('G3:K3')
    sheet.merge_cells('A4:C4')
    sheet.merge_cells('D4:E4')
    sheet.merge_cells('G4:K4')

    sheet['A2'] = "项 目 名 称"
    sheet['D2'] = project.name if project else None
    sheet['F2'] = "项 目 老 板 名 称"
    sheet['G2'] = project.owner if project else None
    sheet['A3'] = "对 账 起 始 日 期"
    sheet['D3'] = start_date
    sheet['F3'] = "对 账 截 止 日 期"
    sheet['G3'] = end_date
    sheet['A4'] = "运 输 单 位 名 称"
    sheet['D4'] = "南平市宏途渣土清运有限公司"
    sheet['F4'] = "公 司 负 责 人"
    sheet['G4'] = "吴 春 才 18905996295"
    
    for row in sheet['A2:K4']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = thin_border

    headers = ["序号", "日期", "", "运输终点", "品类", "车队", "装车方式", "数量", "单位", "终点付费金额", "总金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'B{current_row}:C{current_row}')

    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    total_amount = 0
    for idx, item in enumerate(items, start=1):
        end_site = Site.objects.filter(id=item.endsite_id).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        vehicle = Vehicle.objects.filter(id=item.vehicle_id).first()
        total_price = item.quantity * item.endPayment

        row = [
            idx,
            convert_utc_to_china_time(item.date).split('T')[0],
            "",
            end_site_name,
            goods_name,
            vehicle.license,
            item.get_load_display(),
            item.quantity,
            item.unit,
            item.endPayment,
            total_price
        ]

        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'B{current_row}:C{current_row}')

        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        total_amount += total_price
    if total_amount % 1 == 0:
        total_amount = int(total_amount)
    total_cn = cn2an.an2cn(str(total_amount), "rmb")
    sheet.append(["总 计 金 额", "", "", total_amount, "", "总 计 大 写 (金 额)",total_cn])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:C{current_row}')
    sheet.merge_cells(f'D{current_row}:E{current_row}')
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 运输品类合计
    headers = ["","","序号", "运输终点", "品类", "车队", "合计数量","","单位","终点付费金额","总金额"]
    sheet.append(headers)

    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:H{current_row}')

    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 初步查询，获取所需的基本信息
    initial_query = items.values(
        'endsite_id',
        'goods_id',
        'endPayment',
        'quantity',
        'vehicle_id',
        "unit"
    )

    # 准备聚合结果的字典
    transport_summary_dict = defaultdict(lambda: {
        'endPayment_sum': 0,
        'cost_sum': 0,
        'quantity_sum': 0,
        'vehicle_ids': []
    })

    # 循环处理每个初步查询的结果
    for item in initial_query:
        key = (item['endsite_id'], item['goods_id'], item['endPayment'], item['unit'])
        transport_summary_dict[key]['endPayment_sum'] += item['endPayment']
        transport_summary_dict[key]['cost_sum'] += item['quantity'] * item['endPayment']
        transport_summary_dict[key]['quantity_sum'] += item['quantity']
        transport_summary_dict[key]['vehicle_ids'].append(item['vehicle_id'])

    # 将聚合结果转换为列表
    transport_summary = []
    for (endsite_id, goods_id, endPayment, unit), data in transport_summary_dict.items():
        transport_summary.append({
            'endsite_id': endsite_id,
            'goods_id': goods_id,
            'endPayment_sum': data['endPayment_sum'],
            'quantity_sum': data['quantity_sum'],
            'cost_sum': data['cost_sum'],
            'vehicle_ids': data['vehicle_ids'],
            'endPayment': endPayment,
            'unit': unit
        })


    row1 = sheet.max_row    
    total_sum1 = 0.0
    total_sum2 = 0.0
    for idx, item in enumerate(transport_summary, start=1):
        end_site = Site.objects.filter(id=item['endsite_id']).first()
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item['goods_id']).first()
        goods_name = goods.name if goods else "无"
        vehicle_ids = item['vehicle_ids'][:3]  # Extract up to the first three vehicle IDs
        vehicles = Vehicle.objects.filter(id__in=vehicle_ids)
        vehicle_names ='，'.join([vehicle.license for vehicle in vehicles])
        # Accumulate sums
        total_sum1 += item['cost_sum']
        row = [
            "",
            "",
            idx,
            end_site_name,
            goods_name,
            vehicle_names,
            item['quantity_sum'],
            "",
            item['unit'],
            item['endPayment'],
            item['cost_sum'],
        ]
        sheet.append(row)
        current_row = sheet.max_row
        sheet.merge_cells(f'G{current_row}:H{current_row}')
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border


    sheet.append(["","","合计","-","-","-",total_sum1,"","","",""])
    current_row = sheet.max_row
    sheet.merge_cells(f'G{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row2 = sheet.max_row
    sheet.merge_cells(start_row=row1, start_column=1, end_row=row2, end_column=2)
    cell = sheet.cell(row=row1, column=1)
    cell.value = "运输品类合计"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

    sheet.append(["工 地 负 责 人（ 签 字 确 认 ) ：","","","","","运 输 单 位 负 责 人 (  签 字 确 认 ) ："])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:E{current_row}')
    sheet.merge_cells(f'F{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
    
    sheet.append(["经 营 范 围 ： 建筑垃圾清运，砂石料运输及销售，供应铺路石渣，云梯车租赁。"])
    current_row = sheet.max_row
    sheet.merge_cells(f'A{current_row}:K{current_row}')
    for cell in sheet[current_row]:
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    sheet.append(["立 信 于 心 ， 尽 责 至 善！"])
    centered_row = sheet.max_row
    sheet.merge_cells(f'A{centered_row}:K{centered_row}')
    for cell in sheet[centered_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=16, bold=True)
        cell.border = thin_border

    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 18.5
        for cell in row:
            if not cell.font:
                set_font_and_alignment(cell)
    sheet.row_dimensions[1].height = 25  
    sheet.row_dimensions[2].height = 25  
    sheet.row_dimensions[3].height = 25  
    sheet.row_dimensions[sheet.max_row].height = 22  
    sheet.row_dimensions[sheet.max_row-2].height = 30  

    sheet.page_setup.scale = 53

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_xlsx:
        workbook.save(tmp_xlsx.name)
        tmp_xlsx_path = tmp_xlsx.name
    
    pdf_path = tmp_xlsx_path.replace('.xlsx', '.pdf')

    subprocess.run(['unoconv', '-f', 'pdf', tmp_xlsx_path])

    with open(pdf_path, 'rb') as pdf_file:
        response = HttpResponse(pdf_file.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment;filename=transport_statement.pdf'

    os.remove(tmp_xlsx_path)
    os.remove(pdf_path)

    return response


def detail_excel(req:HttpRequest):
    failure_response, user = get_user_from_request(req,'POST')
    if failure_response:
        return failure_response
    body = json.loads(req.body.decode("utf-8"))
    item_ids = require(body, "item_ids", "list", err_msg="Missing or error type of [item_ids]")
    # 创建Excel工作簿
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "宏途清运明细表"
    default_font = Font(size=10)
    def set_font_and_alignment(cell):
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # 设置列宽
    column_widths = {
        'A': 8.05,    # 序号
        'B': 13.72,   # 日期
        'C': 11.05,   # 项目
        'D': 9.38,   # 老板
        'E': 11.72,   # 起点工地
        'F': 11.72,   # 终点工地
        'G': 13.72,   # 车牌
        'H': 9.38,   # 司机
        'I': 11.72,   # 装车方式
        'J': 8.05,   # 品类
        'K': 8.05,   # 单位
        'L': 8.05,   # 数量
        'M': 15.38,   # 工地承接单价
        'N': 15.38,   # 起点补贴金额
        'O': 15.38,   # 弃点付费金额
        'P': 15.38,   # 终点付费金额
        'Q': 13.55,   # 给司机单价
        'R': 20.05,   # 备注
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # 添加表头
    sheet.merge_cells('A1:R1')
    title_cell = sheet['A1']
    title_cell.value = "宏途清运明细表"
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=16, bold=True)
    current_row = sheet.max_row
    sheet.row_dimensions[current_row].height = 28.5
    # 列标题
    headers = ["序号", "日期", "项目", "老板","运输起点", "运输终点", "车牌","司机", "装车方式","品类","单位","数量",  "工地承接单价", "起点补贴金额", "弃点付费金额","终点付费金额","给司机单价","备注"]
    sheet.append(headers)
    
    current_row = sheet.max_row    
    for cell in sheet[current_row]:
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[current_row].height = 22.5
    # 获取过滤的Item
    items = Item.objects.filter(id__in=item_ids, if_delete=False)
    
    for idx, item in enumerate(items, start=1):
        # get necessary info
        project = Project.objects.filter(id=item.project_id).first()
        project_name = project.name if project else "无"
        project_owner = project.owner if project else "无"
        start_site = Site.objects.filter(id=item.startsite_id).first()
        end_site = Site.objects.filter(id=item.endsite_id).first()
        start_site_name = start_site.name if start_site else "无"
        end_site_name = end_site.name if end_site else "无"
        goods = Goods.objects.filter(id=item.goods_id).first()
        goods_name = goods.name if goods else "无"
        vehicle = Vehicle.objects.filter(id=item.vehicle_id).first()
        driver = vehicle.driver if vehicle else "无"
        license = vehicle.license if vehicle else "无"
        row = [
            idx,
            convert_utc_to_china_time(item.date).split('T')[0],
            project_name,
            project_owner,
            start_site_name,
            end_site_name,
            license,
            driver,
            item.get_load_display(),
            goods_name,
            item.unit,
            item.quantity,
            item.contractorPrice,
            item.startSubsidy,
            item.endSubsidy,
            item.endPayment,
            item.driverPrice,
            item.note
            ]
        # 将数据追加到sheet中，并合并相应的单元格
        sheet.append(row)
        current_row = sheet.max_row
        for cell in sheet[current_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=10)
        sheet.row_dimensions[current_row].height = 22.5  # 设置数据行高
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    response = HttpResponse(file_stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment;filename=transport_statement.xlsx'
    return response
